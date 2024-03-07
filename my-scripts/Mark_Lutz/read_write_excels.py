from openpyxl import load_workbook
excel_file = 'sample_config.xlsx'

workbook = load_workbook(filename = excel_file)
sheet = workbook.active

f = sheet.rows
row_num,col_num = 0,0

for row in f:
    row_num += 1

    header = row[0].value ##First column val

    if header and header.startswith('INGESTION'):
        if 'my_col' in [x.value.lower() for x in row if x.value]:
            continue
        else:
            row_num_of_col = len([x for x in row if x.value])
            sheet.cell(row=row_num,column=row_num_of_col+1).value = 'my_new_col'
        
            for ingest_rows in f:
                row_num += 1
                if not ingest_rows[0].value and not ingest_rows[1].value:
                    break
                sheet.cell(row=row_num,column=row_num_of_col+1).value = 'my_new_val'

workbook.save(filename = excel_file)





